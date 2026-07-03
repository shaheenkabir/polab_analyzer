#!/usr/bin/env python3
import argparse
import os
from posixpath import abspath
import re
import sys
import warnings
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from scipy.signal import find_peaks, savgol_filter

warnings.simplefilter(action="ignore", category=FutureWarning)


# --- Helper: test run without saving ---
def test_parameters(excel_file, lower_lim, upper_lim, distance, prominence, height):
    discarded_sheets = 0
    total_sheets = 0

    try:
        xl = pd.ExcelFile(excel_file)
        # Process sheets that match the 's' followed by digits pattern
        sheet_names = [s for s in xl.sheet_names if re.match(r"^s\d+$", s)]
    except Exception:
        return 0, 0

    for sheet_name in sheet_names:
        try:
            data = pd.read_excel(excel_file, sheet_name=sheet_name)
            length = data.iloc[:, 0]
            intensity = data.iloc[:, 1]

            norm_intensity = intensity / intensity.max()
            inverted_intensity = 1 - norm_intensity
            norm_length = length / length.max()
            percent_length = norm_length * 100

            smoothed = savgol_filter(inverted_intensity, 11, 2)
            diff = np.diff(smoothed, prepend=smoothed[0])
            change_points = np.diff(np.sign(diff), prepend=0)
            change_point_flags = (change_points < 0).astype(int)

            peaks, _ = find_peaks(
                smoothed, distance=distance, prominence=prominence, height=height
            )
            peak_percent_lengths = percent_length.iloc[peaks].values
            valid_peak_mask = (peak_percent_lengths > lower_lim) & (
                    peak_percent_lengths < upper_lim
            )
            peak_percent_lengths = peak_percent_lengths[valid_peak_mask]

            data["Percent Length"] = percent_length
            data["Change Point"] = change_point_flags

            change_df = data[
                (data["Change Point"] == 1) & (data["Percent Length"] < upper_lim)
                ]
            closest_matches = []
            for px in peak_percent_lengths:
                if not change_df.empty:
                    closest = change_df.iloc[
                        (change_df["Percent Length"] - px).abs().argsort()[:1]
                    ]
                    closest_val = closest["Percent Length"].values[0]
                    closest_matches.append(closest_val)

            total_sheets += 1
            if len(closest_matches) != 7:
                discarded_sheets += 1

        except Exception:
            discarded_sheets += 1

    return discarded_sheets, total_sheets



def test_parameters2(input_folder, lower_lim, upper_lim, distance, prominence, height):
    discarded_files = 0
    total_files = 0

    for filename in os.listdir(input_folder):
        if filename.endswith(".xlsx"):
            input_path = os.path.join(input_folder, filename)
            try:
                data = pd.read_excel(input_path)
                length = data.iloc[:, 0]
                intensity = data.iloc[:, 1]

                norm_intensity = intensity / intensity.max()
                inverted_intensity = 1 - norm_intensity
                norm_length = length / length.max()
                percent_length = norm_length * 100

                smoothed = savgol_filter(inverted_intensity, 11, 2)
                diff = np.diff(smoothed, prepend=smoothed[0])
                change_points = np.diff(np.sign(diff), prepend=0)
                change_point_flags = (change_points < 0).astype(int)

                peaks, _ = find_peaks(
                    smoothed, distance=distance, prominence=prominence, height=height
                )
                peak_percent_lengths = percent_length.iloc[peaks].values
                valid_peak_mask = (peak_percent_lengths > lower_lim) & (
                        peak_percent_lengths < upper_lim
                )
                peak_percent_lengths = peak_percent_lengths[valid_peak_mask]

                data["Percent Length"] = percent_length
                data["Change Point"] = change_point_flags

                change_df = data[
                    (data["Change Point"] == 1) & (data["Percent Length"] < upper_lim)
                    ]
                closest_matches = []
                for px in peak_percent_lengths:
                    if not change_df.empty:
                        closest = change_df.iloc[
                            (change_df["Percent Length"] - px).abs().argsort()[:1]
                        ]
                        closest_val = closest["Percent Length"].values[0]
                        closest_matches.append(closest_val)

                total_files += 1
                if len(closest_matches) != 7:
                    discarded_files += 1

            except Exception:
                discarded_files += 1

    return discarded_files, total_files  


def generate_stripe_summary(output_folder):
    summary_output_path = os.path.join(output_folder, "stripe_summary.xlsx")

    all_rows = []
    processed_files = sorted(
        [
            f
            for f in os.listdir(output_folder)
            if f.startswith("processed_") and f.endswith(".xlsx")
        ],
        key=lambda x: (
            int(re.findall(r"(\d+)", x)[-1]) if re.findall(r"(\d+)", x) else 0
        ),
    )

    for fname in processed_files:
        fpath = os.path.join(output_folder, fname)
        try:
            wb = load_workbook(fpath, data_only=True)
            if "Peaks" not in wb.sheetnames:
                continue

            peaks_ws = wb["Peaks"]
            peaks = []
            for i, r in enumerate(
                    peaks_ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                if r[0] is not None:
                    peaks.append(r[0])

            if len(peaks) != 7:
                print(f"\033[91mSkipping {fname}\033[0m: found {len(peaks)} stripes, not 7.")
                continue

            row_data = {"File": fname}
            for i, p in enumerate(peaks, start=1):
                row_data[f"Stripe-{i}"] = p

            all_rows.append(row_data)

        except Exception as e:
            print(f"Error reading {fname}: {e}")

    # Save summary
    summary_df = pd.DataFrame(all_rows)
    summary_df.to_excel(summary_output_path, index=False)
    print(f"\033[92mStripe summary written to: {summary_output_path}\033[0m")


def run_pipeline(
    excel_file,
    lower_lim,
    upper_lim,
    test,
    distance,
    prominence,
    height,
    distances,
    prom_range,
    height_range,
):
    """Auto-tune (or use fixed) peak-detection parameters and process every
    s1, s2, s3, ... sheet of a single Excel file. Shared by both --file mode
    (one file) and --folder mode (looped over every xlsx file in the folder)."""
    if test:
        # --- Grid search mode ---
        distances_list = distances
        prominences = np.arange(
            prom_range[0],
            prom_range[1] + prom_range[2],
            prom_range[2],
        )
        heights = np.arange(
            height_range[0],
            height_range[1] + height_range[2],
            height_range[2],
        )

        best_params = None
        best_discarded = float("inf")

        total_iterations = len(distances_list) * len(prominences) * len(heights)
        current_iteration = 0

        print(f"\n\033[94mRunning parameter grid search over sheets in '{os.path.basename(excel_file)}'...\033[0m")

        for d in distances_list:
            for p in prominences:
                for h in heights:
                    discarded, total = test_parameters(
                        excel_file, lower_lim, upper_lim, d, p, h
                    )
                    if discarded < best_discarded:
                        best_discarded = discarded
                        best_params = (d, p, h)

                    current_iteration += 1
                    percent = (current_iteration / total_iterations) * 100

                    sys.stdout.write(
                        f"\r\033[92mProgress: [{percent:3.0f}%]\033[0m Processing parameter set {current_iteration}/{total_iterations}..."
                    )
                    sys.stdout.flush()

        print("\n")

        print(
            f"\033[96mFound best parameters:\033[0m distance={best_params[0]}, "
            f"prominence={best_params[1]:.2f}, height={best_params[2]:.2f} "
            f"(\033[91mdiscarded {best_discarded} sheets\033[0m)"
        )

        final_distance, final_prominence, final_height = best_params

    else:
        # --- Direct mode ---
        final_distance = distance
        final_prominence = prominence
        final_height = height

        print(
            f"\n Using user-specified parameters: "
            f"distance={final_distance}, prominence={final_prominence}, height={final_height}"
        )

    process_excel_file(
        excel_file,
        lower_lim,
        upper_lim,
        distance=final_distance,
        prominence=final_prominence,
        height=final_height,
    )
    
    
def run_pipeline2(input_folder, lower_lim, upper_lim, test, distance, prominence, height, distances, prom_range, height_range):
    if test:
        # --- Grid search mode ---
        distances_list = distances
        prominences = np.arange(
            prom_range[0],
            prom_range[1] + prom_range[2],
            prom_range[2],
        )
        heights = np.arange(
            height_range[0],
            height_range[1] + height_range[2],
            height_range[2],
        )

        best_params = None
        best_discarded = float("inf")

        total_iterations = len(distances_list) * len(prominences) * len(heights)
        current_iteration = 0

        print(f"\n\033[94mRunning parameter grid search over files in '{os.path.basename(input_folder)}'...\033[0m")

        for d in distances_list:
            for p in prominences:
                for h in heights:
                    discarded, total = test_parameters2(
                        input_folder, lower_lim, upper_lim, d, p, h
                    )
                    if discarded < best_discarded:
                        best_discarded = discarded
                        best_params = (d, p, h)

                    current_iteration += 1
                    percent = (current_iteration / total_iterations) * 100

                    sys.stdout.write(
                        f"\r\033[92mProgress: [{percent:3.0f}%]\033[0m Processing parameter set {current_iteration}/{total_iterations}..."
                    )
                    sys.stdout.flush()

        print("\n")

        print(
            f"\033[96mFound best parameters:\033[0m distance={best_params[0]}, "
            f"prominence={best_params[1]:.2f}, height={best_params[2]:.2f} "
            f"(\033[91mdiscarded {best_discarded} files\033[0m)"
        )

        final_distance, final_prominence, final_height = best_params

    else:
        # --- Direct mode ---
        final_distance = distance
        final_prominence = prominence
        final_height = height
        print(
            f"\n Using user-specified parameters: "
            f"distance={final_distance}, prominence={final_prominence}, height={final_height}"
        )

    process_folder(
        input_folder,
        lower_lim,
        upper_lim,
        distance=final_distance,
        prominence=final_prominence,
        height=final_height,
    )


def process_folder(input_folder, lower_lim, upper_lim, distance=10, prominence=0.03, height = 0.1):
    output_folder = os.path.join(input_folder, "results")
    os.makedirs(output_folder, exist_ok=True)

    for filename in os.listdir(input_folder):
        if filename.endswith(".xlsx"):
            input_path = os.path.join(input_folder, filename)
            output_path = os.path.join(output_folder, f"processed_{filename}")
            temp_plot_path = "temp_plot.png"
            try:
                data = pd.read_excel(input_path)
                length = data.iloc[:, 0]
                intensity = data.iloc[:, 1]

                norm_intensity = intensity / intensity.max()
                inverted_intensity = 1 - norm_intensity
                norm_length = length / length.max()
                percent_length = norm_length * 100

                smoothed = savgol_filter(inverted_intensity, 11, 2)
                diff = np.diff(smoothed, prepend=smoothed[0])
                change_points = np.diff(np.sign(diff), prepend=0)
                change_point_flags = (change_points < 0).astype(int)
                
                # Detect peaks
                peaks, _ = find_peaks(smoothed, distance=distance, prominence=prominence, height=height)
                peak_percent_lengths = percent_length.iloc[peaks].values
                peak_values = smoothed[peaks]
                valid_peak_mask = (peak_percent_lengths > lower_lim) & (peak_percent_lengths < upper_lim)
                peak_percent_lengths = peak_percent_lengths[valid_peak_mask]
                peak_values = peak_values[valid_peak_mask]

                data['Normalized Intensity'] = norm_intensity
                data['Inverted Intensity'] = inverted_intensity
                data['Normalized Length'] = norm_length
                data['Percent Length'] = percent_length
                data['Smoothed'] = smoothed
                data['Difference'] = diff
                data['Change Point'] = change_point_flags
                data.to_excel(output_path, index=False)

                plt.figure(figsize=(5, 3))
                plt.plot(percent_length, smoothed, label="Smoothed", color='blue')
                plt.scatter(peak_percent_lengths, peak_values, color='red', s=50, label=f"Peaks > {lower_lim}%")
                plt.title("Smoothed Inverted Intensity")
                plt.xlabel("Percent Length (%)")
                plt.ylabel("Inverted Intensity")
                plt.legend()
                plt.tight_layout()
                plt.savefig(temp_plot_path, dpi=200)
                plt.close()
                wb = load_workbook(output_path)
                ws = wb.active
                img = XLImage(temp_plot_path)
                img.width = 360
                img.height = 220
                ws.add_image(img, "K2")
                wb.save(output_path)

                # Add Peaks Sheets
                if "Peaks" in wb.sheetnames:
                    del wb["Peaks"]
                peak_ws = wb.create_sheet("Peaks")
                peak_ws.append(["Percent Length", "Inverted Intensity"])
                for x, y in zip(peak_percent_lengths, peak_values):
                    peak_ws.append([x, y])

                # Match peaks with change_df data frame
                change_df = data[(data["Change Point"] == 1) & (data["Percent Length"] > lower_lim) & (
                            data["Percent Length"] < upper_lim)]
                closest_matches = []
                for px in peak_percent_lengths:
                    if not change_df.empty:
                        closest = change_df.iloc[(change_df["Percent Length"] - px).abs().argsort()[:1]]
                        closest_val = closest["Percent Length"].values[0]
                        closest_matches.append(closest_val)
                        
                # Discard files which have not exactly 7 stripes.
                if len(closest_matches) != 7:
                    print(
                        f"\033[91mSkipping {filename}:\033[0m found {len(closest_matches)} stripes, not exactly 7."
                        )
                    if os.path.exists(temp_plot_path):
                        os.remove(temp_plot_path)
                    continue
                wb.save(output_path)
                if os.path.exists(temp_plot_path):
                    os.remove(temp_plot_path)

                # Add macro sheets.
                macro_df = pd.DataFrame({
                    "Percent Length": percent_length,
                    "Inverted Intensity": inverted_intensity
                })
                with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
                    macro_df.to_excel(writer, sheet_name="Macro", index=False)

                # Add macro 600 points sheet
                indices = np.linspace(0, len(macro_df) - 1, 600, dtype=int)
                squeezed_macro_df = macro_df.iloc[indices].reset_index(drop=True)
                with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    squeezed_macro_df.to_excel(writer, sheet_name='Macro_600points', index=False)
            except Exception as e:
                print(f"Error processing {filename}: {e}")

    def extract_numeric_suffix(filename):
        nums = re.findall(r'(\d+)', filename)
        return tuple(map(int, nums)) if nums else (0,)
    # Combine all Macro_600points into one Excel
    macro_percent_cols = []
    macro_intensity_cols = []
    sample_names = []
    processed_files = sorted(
        [f for f in os.listdir(output_folder) if f.startswith("processed_") and f.endswith(".xlsx")],
        key=lambda x: int(re.findall(r'(\d+)', x)[-1])
    )
    for idx, fname in enumerate(processed_files, start=1):
        fpath = os.path.join(output_folder, fname)
        try:
            df = pd.read_excel(fpath, sheet_name="Macro")
            percent_col = df.iloc[:, 0] / df.iloc[:, 0].max()
            intensity_col = df.iloc[:, 1]
            indices = np.linspace(0, len(df) - 1, 600, dtype=int)
            percent_squeezed = percent_col.iloc[indices].reset_index(drop=True)
            intensity_squeezed = intensity_col.iloc[indices].reset_index(drop=True)
            
            label = os.path.basename(input_folder) + f"_{idx}"
            sample_names.append(label)
            macro_percent_cols.append(percent_squeezed.rename(label))
            macro_intensity_cols.append(intensity_squeezed.rename(label))
        except Exception:
            pass
            
    percent_df = pd.concat(macro_percent_cols, axis=1)
    intensity_df = pd.concat(macro_intensity_cols, axis=1)
    percent_df[""] = ""
    intensity_df[""] = ""
    percent_df["Average"] = percent_df.select_dtypes(include=[np.number]).mean(axis=1)
    intensity_df["Average"] = intensity_df.select_dtypes(include=[np.number]).mean(axis=1)
    combined_macro_path = os.path.join(output_folder, "combined_macro_600points.xlsx")
    with pd.ExcelWriter(combined_macro_path) as writer:
        percent_df.to_excel(writer, sheet_name="Percent Length", index=False)
        intensity_df.to_excel(writer, sheet_name="Inverted Intensity", index=False)

    generate_stripe_summary(output_folder)
    
    
    
def process_excel_file(
        excel_file, lower_lim, upper_lim, distance=10, prominence=0.03, height=0.1
):
    parent_folder = os.path.dirname(os.path.abspath(excel_file))
    base_name = os.path.splitext(os.path.basename(excel_file))[0]
    output_folder = os.path.join(parent_folder, "results")
    os.makedirs(output_folder, exist_ok=True)

    xl = pd.ExcelFile(excel_file)
    sheet_names = sorted(
        [s for s in xl.sheet_names if re.match(r"^s\d+$", s)],
        key=lambda x: int(re.findall(r"(\d+)", x)[0])
    )

    for sheet_name in sheet_names:
        output_path = os.path.join(output_folder, f"processed_{base_name}_{sheet_name}.xlsx")
        temp_plot_path = f"temp_plot_{sheet_name}.png"
        try:
            data = pd.read_excel(excel_file, sheet_name=sheet_name)
            length = data.iloc[:, 0]
            intensity = data.iloc[:, 1]

            norm_intensity = intensity / intensity.max()
            inverted_intensity = 1 - norm_intensity
            norm_length = length / length.max()
            percent_length = norm_length * 100

            smoothed = savgol_filter(inverted_intensity, 11, 2)
            diff = np.diff(smoothed, prepend=smoothed[0])
            change_points = np.diff(np.sign(diff), prepend=0)
            change_point_flags = (change_points < 0).astype(int)

            # Detect peaks
            peaks, _ = find_peaks(
                smoothed, distance=distance, prominence=prominence, height=height
            )
            peak_percent_lengths = percent_length.iloc[peaks].values
            peak_values = smoothed[peaks]
            valid_peak_mask = (peak_percent_lengths > lower_lim) & (
                    peak_percent_lengths < upper_lim
            )
            peak_percent_lengths = peak_percent_lengths[valid_peak_mask]
            peak_values = peak_values[valid_peak_mask]

            data["Normalized Intensity"] = norm_intensity
            data["Inverted Intensity"] = inverted_intensity
            data["Normalized Length"] = norm_length
            data["Percent Length"] = percent_length
            data["Smoothed"] = smoothed
            data["Difference"] = diff
            data["Change Point"] = change_point_flags
            data.to_excel(output_path, index=False)

            plt.figure(figsize=(5, 3))
            plt.plot(percent_length, smoothed, label="Smoothed", color="blue")
            plt.scatter(
                peak_percent_lengths,
                peak_values,
                color="red",
                s=50,
                label=f"Peaks > {lower_lim}%",
            )
            plt.title("Smoothed Inverted Intensity")
            plt.xlabel("Percent Length (%)")
            plt.ylabel("Inverted Intensity")
            plt.legend()
            plt.tight_layout()
            plt.savefig(temp_plot_path, dpi=200)
            plt.close()

            wb = load_workbook(output_path)
            ws = wb.active
            img = XLImage(temp_plot_path)
            img.width = 360
            img.height = 220
            ws.add_image(img, "K2")
            wb.save(output_path)

            # Add Peaks Sheets
            if "Peaks" in wb.sheetnames:
                del wb["Peaks"]
            peak_ws = wb.create_sheet("Peaks")
            peak_ws.append(["Percent Length", "Inverted Intensity"])
            for x, y in zip(peak_percent_lengths, peak_values):
                peak_ws.append([x, y])

            # Match peaks with change_df data frame
            change_df = data[
                (data["Change Point"] == 1)
                & (data["Percent Length"] > lower_lim)
                & (data["Percent Length"] < upper_lim)
                ]
            closest_matches = []
            for px in peak_percent_lengths:
                if not change_df.empty:
                    closest = change_df.iloc[
                        (change_df["Percent Length"] - px).abs().argsort()[:1]
                    ]
                    closest_val = closest["Percent Length"].values[0]
                    closest_matches.append(closest_val)

            # Discard files which have not exactly 7 stripes.
            if len(closest_matches) != 7:
                print(
                    f"\033[91mSkipping sheet {sheet_name}:\033[0m found {len(closest_matches)} stripes, not exactly 7."
                )
                if os.path.exists(temp_plot_path):
                    os.remove(temp_plot_path)
                continue
            wb.save(output_path)
            if os.path.exists(temp_plot_path):
                os.remove(temp_plot_path)

            # Add macro sheets.
            macro_df = pd.DataFrame(
                {
                    "Percent Length": percent_length,
                    "Inverted Intensity": inverted_intensity,
                }
            )
            with pd.ExcelWriter(
                    output_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                macro_df.to_excel(writer, sheet_name="Macro", index=False)

            # Add macro 600 points sheet
            indices = np.linspace(0, len(macro_df) - 1, 600, dtype=int)
            squeezed_macro_df = macro_df.iloc[indices].reset_index(drop=True)
            with pd.ExcelWriter(
                    output_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                squeezed_macro_df.to_excel(
                    writer, sheet_name="Macro_600points", index=False
                )
        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {e}")
            if os.path.exists(temp_plot_path):
                os.remove(temp_plot_path)

    # Combine all Macro_600points into one Excel
    macro_percent_cols = []
    macro_intensity_cols = []
    processed_files = sorted(
        [
            f
            for f in os.listdir(output_folder)
            if f.startswith("processed_") and f.endswith(
            ".xlsx") and "stripe_summary" not in f and "combined_macro" not in f
        ],
        key=lambda x: int(re.findall(r"(\d+)", x)[-1]) if re.findall(r"(\d+)", x) else 0,
    )

    for idx, fname in enumerate(processed_files, start=1):
        fpath = os.path.join(output_folder, fname)
        try:
            df = pd.read_excel(fpath, sheet_name="Macro")
            percent_col = df.iloc[:, 0] / df.iloc[:, 0].max()
            intensity_col = df.iloc[:, 1]
            indices = np.linspace(0, len(df) - 1, 600, dtype=int)
            percent_squeezed = percent_col.iloc[indices].reset_index(drop=True)
            intensity_squeezed = intensity_col.iloc[indices].reset_index(drop=True)

            label = f"{base_name}_{idx}"
            macro_percent_cols.append(percent_squeezed.rename(label))
            macro_intensity_cols.append(intensity_squeezed.rename(label))
        except Exception:
            pass

    if macro_percent_cols and macro_intensity_cols:
        percent_df = pd.concat(macro_percent_cols, axis=1)
        intensity_df = pd.concat(macro_intensity_cols, axis=1)
        percent_df[""] = ""
        intensity_df[""] = ""
        percent_df["Average"] = percent_df.select_dtypes(include=[np.number]).mean(axis=1)
        intensity_df["Average"] = intensity_df.select_dtypes(include=[np.number]).mean(axis=1)

        combined_macro_path = os.path.join(output_folder, "combined_macro_600points.xlsx")
        with pd.ExcelWriter(combined_macro_path) as writer:
            percent_df.to_excel(writer, sheet_name="Percent Length", index=False)
            intensity_df.to_excel(writer, sheet_name="Inverted Intensity", index=False)

    generate_stripe_summary(output_folder)


def main():
    parser = argparse.ArgumentParser(
        description="Embryo stripe analysis: auto-tunes parameters to minimize discarded sheets from a single Excel file."
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--file", "-file", help="Path to a single Excel file containing sheets s1, s2, s3..."
    )
    group.add_argument(
        "--folder",
        "-folder",
        help="Path to a folder containing multiple Excel files, each with sheets s1, s2, s3...",
    )
    parser.add_argument(
        "--lower",
        "-l",
        type=float,
        required=True,
        help="Lower limit for percent length",
    )
    parser.add_argument(
        "--upper",
        "-u",
        type=float,
        required=True,
        help="Upper limit for percent length",
    )

    parser.add_argument(
        "--test",
        type=lambda x: str(x).lower() == "true",
        default=True,
        help="Whether to run parameter testing (true/false, default: true)",
    )
    parser.add_argument(
        "--distance",
        type=int,
        default=15,
        help="Distance for peak detection (used if --test false)",
    )
    parser.add_argument(
        "--prominence",
        type=float,
        default=0.03,
        help="Prominence for peak detection (used if --test false)",
    )
    parser.add_argument(
        "--height",
        type=float,
        default=0.1,
        help="Height for peak detection (used if --test false)",
    )

    # For test mode
    parser.add_argument(
        "--distances",
        nargs="+",
        type=int,
        default=[10, 15, 20, 25, 30, 35],
        help="List of distances to test (default: 10 15 20 25 30 35)",
    )
    parser.add_argument(
        "--prom_range",
        nargs=3,
        type=float,
        default=[0.005, 0.08, 0.05],
        help="Prominence range: start end step (default: 0.05 0.8 0.05)",
    )
    parser.add_argument(
        "--height_range",
        nargs=3,
        type=float,
        default=[0.01, 0.5, 0.05],
        help="Height range: start end step (default: 0.01 0.5 0.05)",
    )

    args = parser.parse_args()
    lower_lim = args.lower
    upper_lim = args.upper

    pipeline_kwargs = dict(
        lower_lim=lower_lim,
        upper_lim=upper_lim,
        test=args.test,
        distance=args.distance,
        prominence=args.prominence,
        height=args.height,
        distances=args.distances,
        prom_range=args.prom_range,
        height_range=args.height_range,
    )

    if args.file:
        # --- Single-file mode: file contains sheets s1, s2, s3... ---
        run_pipeline(args.file, **pipeline_kwargs)

    else:
        # --- Folder mode ---
        input_folder = args.folder

        if not os.path.isdir(input_folder):
            print(f"\033[91mError: '{input_folder}' is not a valid directory.\033[0m")
            sys.exit(1)

        # Simply pass the folder path ONCE to the pipeline instead of looping here
        print(f"\033[94m\nStarting analysis on folder: {abspath(input_folder)}\033[0m")
        run_pipeline2(input_folder, **pipeline_kwargs)

        print("\n Completed !!!")


if __name__ == "__main__":
    main()
