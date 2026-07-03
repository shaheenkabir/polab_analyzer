#!/usr/bin/env python3
import argparse
import os
import re
import sys
import warnings
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from scipy.signal import savgol_filter, find_peaks

warnings.simplefilter(action="ignore", category=FutureWarning)

def analyze_intensity_profile(x_orig, y_orig, sample_name):
    # Normalize length to a 0-100 scale
    if (x_orig.max() - x_orig.min()) == 0:
        print(f"  - Skipping {sample_name}: Length data is constant.")
        return None
    x_norm = 100 * (x_orig - x_orig.min()) / (x_orig.max() - x_orig.min())

    # Normalize intensity to 0-1 and then invert it (1 - normalized_value)
    if (y_orig.max() - y_orig.min()) == 0:
        print(f"  - Skipping {sample_name}: Intensity data is constant.")
        return None
    y_norm = (y_orig - y_orig.min()) / (y_orig.max() - y_orig.min())
    y_proc = 1 - y_norm

    # Smooth the PROCESSED data
    if len(y_proc) > 51:
        window_length = 51
    else:
        window_length = max(5, len(y_proc) // 2 * 2 + 1)

    if window_length <= 3:
        print(f"  - Skipping {sample_name}: Not enough data points to process.")
        return None

    polyorder = 3
    y_smooth = savgol_filter(y_proc, window_length, polyorder)

    # Find peaks on the PROCESSED data
    peaks, _ = find_peaks(y_smooth, prominence=0.05, height=0.05)

    change_point_x = None
    peak_indices = []

    # Applying the peak finding logic.
    if len(peaks) >= 2:
        # --- Two-Peak Method ---
        peak_prominences = y_smooth[peaks]
        top_two_indices = np.argsort(peak_prominences)[-2:]
        peak_indices = sorted(peaks[top_two_indices])
        first_peak_idx, second_peak_idx = peak_indices[0], peak_indices[1]

        section = slice(first_peak_idx, second_peak_idx + 1)
        if len(y_smooth[section]) > 1:
            intensity_gradient = np.gradient(y_smooth[section])
            decreasing_indices = np.where(intensity_gradient < -0.002)[0]
            if len(decreasing_indices) > 0:
                groups = np.split(decreasing_indices, np.where(np.diff(decreasing_indices) > 1)[0] + 1)
                longest_group = max(groups, key=len)
                midpoint_local_idx = (longest_group[0] + longest_group[-1]) // 2
                change_point_x = x_norm[section][midpoint_local_idx]
    else:
        # --- Single-Peak Method ---
        first_peak_idx = np.argmax(y_smooth) if len(peaks) == 0 else peaks[0]
        peak_indices = [first_peak_idx]

        section = slice(first_peak_idx, len(y_smooth))
        if len(y_smooth[section]) > 1:
            intensity_gradient = np.gradient(y_smooth[section])
            steepest_decrease_local_idx = np.argmin(intensity_gradient)
            change_point_x = x_norm[section][steepest_decrease_local_idx]

    return {
        "x_norm": x_norm,
        "y_proc": y_proc,
        "y_smooth": y_smooth,
        "peak_indices": peak_indices,
        "change_point_x": change_point_x,
    }


def _natural_sort_key(s):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", s)]


def generate_midpoint_summary(output_folder):
    summary_output_path = os.path.join(output_folder, "midpoint_summary.xlsx")

    all_rows = []
    processed_files = sorted(
        [
            f
            for f in os.listdir(output_folder)
            if f.startswith("processed_") and f.endswith(".xlsx")
        ],
        key=_natural_sort_key,
    )

    for fname in processed_files:
        fpath = os.path.join(output_folder, fname)
        try:
            wb = load_workbook(fpath, data_only=True)
            if "Midpoint" not in wb.sheetnames:
                continue

            mp_ws = wb["Midpoint"]
            rows = list(mp_ws.iter_rows(min_row=2, max_row=2, values_only=True))
            if not rows:
                continue

            sample, midpoint = rows[0]
            all_rows.append(
                {
                    "Sample": sample,
                    "Midpoint (% Egg Length)": midpoint if midpoint is not None else "N/A",
                }
            )

        except Exception as e:
            print(f"Error reading {fname}: {e}")

    summary_df = pd.DataFrame(all_rows)
    summary_df.to_excel(summary_output_path, index=False)
    print(f"\033[92mMidpoint summary written to: {summary_output_path} \033[0m")


def process_dataset(x_orig, y_orig, sample_name, output_path):
    """Run the analysis/plot/save pipeline for a single (x, y) profile and
    write the result to output_path. Shared by both --folder mode (one xlsx
    file = one sample) and --file mode (one sheet = one sample)."""
    base_name = os.path.splitext(os.path.basename(output_path))[0]
    temp_plot_path = os.path.join(
        os.path.dirname(output_path), f"temp_plot_{base_name}.png"
    )

    result = analyze_intensity_profile(x_orig, y_orig, sample_name)
    if result is None:
        return None

    x_norm = result["x_norm"]
    y_proc = result["y_proc"]
    y_smooth = result["y_smooth"]
    peak_indices = result["peak_indices"]
    change_point_x = result["change_point_x"]

    try:
        # Save processed data
        out_df = pd.DataFrame(
            {
                "Length (orig)": x_orig,
                "Intensity (orig)": y_orig,
                "Percent Length": x_norm,
                "Processed Intensity": y_proc,
                "Smoothed Intensity": y_smooth,
            }
        )
        out_df.to_excel(output_path, index=False)

        # Plot the results using PROCESSED data.
        plt.figure(figsize=(12, 7))
        plt.plot(x_norm, y_proc, color="grey", alpha=0.6, label="Raw datapoints")
        plt.plot(x_norm, y_smooth, color="black", linewidth=2, label="Smoothed Line")

        if len(peak_indices) > 0:
            plt.axvline(
                x=x_norm[peak_indices[0]],
                color="blue",
                linestyle="--",
                label=f"Peak 1 (at {x_norm[peak_indices[0]]:.2f})",
            )
        if len(peak_indices) > 1:
            plt.axvline(
                x=x_norm[peak_indices[1]],
                color="green",
                linestyle="--",
                label=f"Peak 2 (at {x_norm[peak_indices[1]]:.2f})",
            )

        if change_point_x is not None:
            plt.axvline(
                x=change_point_x,
                color="red",
                linestyle="--",
                linewidth=2,
                label="Midpoint of Decrease",
            )
            annotation_text = f"Midpoint: {change_point_x:.2f}"
            y_range = plt.ylim()[1] - plt.ylim()[0]
            plt.text(
                change_point_x + 2,
                plt.ylim()[0] + y_range * 0.5,
                annotation_text,
                color="red",
                fontsize=12,
                bbox=dict(facecolor="white", alpha=0.8, edgecolor="red"),
            )

        plt.title(f"Intensity Profile for: {sample_name}", fontsize=16)
        plt.xlabel("Embryo Length (%)", fontsize=12)
        plt.ylabel("Intensity", fontsize=12)
        plt.grid(True, which="both", linestyle="--", linewidth=0.5)
        plt.legend()
        plt.tight_layout()
        plt.savefig(temp_plot_path, dpi=200)
        plt.close()

        # Embed the plot image into the processed xlsx.
        wb = load_workbook(output_path)
        ws = wb.active
        img = XLImage(temp_plot_path)
        img.width = 480
        img.height = 280
        ws.add_image(img, "H2")

        if "Midpoint" in wb.sheetnames:
            del wb["Midpoint"]
        mp_ws = wb.create_sheet("Midpoint")
        mp_ws.append(["Sample", "Midpoint (% Egg Length)"])
        mp_ws.append([sample_name, change_point_x if change_point_x is not None else "N/A"])

        wb.save(output_path)

    except Exception as e:
        print(f"  - Error processing {sample_name}: {e}")
        return None
    finally:
        if os.path.exists(temp_plot_path):
            os.remove(temp_plot_path)

    return change_point_x


def process_excel_file(excel_file, output_folder):
    """--folder mode: one xlsx file = one sample."""
    base_name = os.path.splitext(os.path.basename(excel_file))[0]
    output_path = os.path.join(output_folder, f"processed_{base_name}.xlsx")

    try:
        data = pd.read_excel(excel_file)
        x_orig = data.iloc[:, 0].values
        y_orig = data.iloc[:, 1].values
    except Exception as e:
        print(f"  - Error reading {os.path.basename(excel_file)}: {e}")
        return None

    return process_dataset(x_orig, y_orig, base_name, output_path)


def process_file_sheets(excel_file, output_folder):
    """--file mode: one xlsx file containing sheets s1, s2, s3, ... = multiple samples."""
    base_name = os.path.splitext(os.path.basename(excel_file))[0]

    try:
        xl = pd.ExcelFile(excel_file)
        sheet_names = sorted(
            [s for s in xl.sheet_names if re.match(r"^s\d+$", s, re.IGNORECASE)],
            key=lambda s: int(re.findall(r"\d+", s)[0]),
        )
    except Exception as e:
        print(f"\033[91mError reading '{excel_file}': {e}\033[0m")
        return

    if not sheet_names:
        print(
            f"\033[91mNo sheets matching the 's1', 's2', ... pattern were found in "
            f"'{excel_file}'.\033[0m"
        )
        return

    total_sheets = len(sheet_names)
    print(
        f"\033[96m\nFound {total_sheets} sheets in '{os.path.basename(excel_file)}'.\033[0m "
        f"\n\033[91mStarting analysis...\033[0m"
    )

    for i, sheet_name in enumerate(sheet_names, start=1):
        percent = (i / total_sheets) * 100
        sys.stdout.write(
            f"\r\033[92mProgress: [{percent:3.0f}%]\033[0m Processing sheets {i}/{total_sheets}..."
        )
        sys.stdout.flush()

        sample_name = f"{base_name}_{sheet_name}"
        output_path = os.path.join(output_folder, f"processed_{sample_name}.xlsx")

        try:
            data = pd.read_excel(excel_file, sheet_name=sheet_name)
            x_orig = data.iloc[:, 0].values
            y_orig = data.iloc[:, 1].values
        except Exception as e:
            print(f"\n  - Error reading sheet {sheet_name}: {e}")
            continue

        process_dataset(x_orig, y_orig, sample_name, output_path)

    print()


def main():
    parser = argparse.ArgumentParser(
        description="Hb fluorescence intensity profile batch analyzer: finds the midpoint "
        "of the intensity decrease for every sample file in a folder."
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--folder",
        "-folder",
        help="Path to a folder containing input .xlsx sample files (one sample per file)",
    )
    group.add_argument(
        "--file",
        "-file",
        help="Path to a single .xlsx file containing sheets s1, s2, s3, ... "
        "(one sample per sheet)",
    )
    args = parser.parse_args()

    if args.folder:
        input_folder = args.folder

        if not os.path.isdir(input_folder):
            print(f"\033[91mError: '{input_folder}' is not a valid directory.\033[0m")
            sys.exit(1)

        xlsx_files = sorted(
            f
            for f in os.listdir(input_folder)
            if f.endswith(".xlsx") and not f.startswith("processed_") and "summary" not in f.lower()
        )

        if not xlsx_files:
            print(f"\nNo .xlsx files found in '{input_folder}'. Exiting.")
            sys.exit(0)

        output_folder = os.path.join(input_folder, "results")
        os.makedirs(output_folder, exist_ok=True)

        print(f"\033[96m\nFound total {len(xlsx_files)} Excel files.\033[0m \n\033[91mStarting analysis...\033[0m")

        total_files = len(xlsx_files)
        for i, fname in enumerate(xlsx_files, start=1):
            fpath = os.path.join(input_folder, fname)
            percent = (i / total_files) * 100

            sys.stdout.write(
                f"\r\033[92mProgress: [{percent:3.0f}%]\033[0m Processing Excel Files {i}/{total_files}..."
            )
            sys.stdout.flush()

            process_excel_file(fpath, output_folder)

        print("\n Completed !!!")

    else:
        input_file = args.file

        if not os.path.isfile(input_file):
            print(f"\033[91mError: '{input_file}' is not a valid file.\033[0m")
            sys.exit(1)

        output_folder = os.path.join(os.path.dirname(os.path.abspath(input_file)), "results")
        os.makedirs(output_folder, exist_ok=True)

        process_file_sheets(input_file, output_folder)

        print("\n Completed !!!")

    generate_midpoint_summary(output_folder)

if __name__ == "__main__":
    main()
